import xml.etree.ElementTree as ET
import os
import shutil
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook


def get_name_from_xml(xml_file):
    try:
        print(f"Parsing XML file: {xml_file}")
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Print the XML structure for debugging
        print(f"XML structure:\n{ET.tostring(root, encoding='unicode')}")

        # Try to find <NamCE17> first, then <NamEX17>
        name_element = root.find('.//NamCE17')
        if name_element is None:
            name_element = root.find('.//NamEX17')

        if name_element is None:
            raise ValueError(f"No <NamCE17> or <NamEX17> element found in the XML file: {xml_file}")

        name = name_element.text.strip()
        print(f"Extracted name: {name}")
        return name
    except ET.ParseError as e:
        print(f"Error parsing XML file {xml_file}: {e}")
        raise
    except Exception as e:
        print(f"An error occurred while processing XML file {xml_file}: {e}")
        raise


def extract_additional_data_from_xml(xml_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Extract "τελωνειο" from <MesRecMeS6>
        teloneio_element = root.find('.//MesRecMES6')
        if teloneio_element is None:
            raise ValueError(f"No <MesRecMES6> element found in the XML file: {xml_file}")
        teloneio = teloneio_element.text.strip()
        print(f"Extracted τελωνειο: {teloneio}")

        # Extract "αριθμος τελωνειου" from <RefNumHEA4>
        ref_num_element = root.find('.//RefNumHEA4')
        if ref_num_element is None:
            raise ValueError(f"No <RefNumHEA4> element found in the XML file: {xml_file}")
        ref_num_parts = ref_num_element.text.split('/')
        if len(ref_num_parts) < 3:
            raise ValueError(f"Invalid <RefNumHEA4> format in the XML file: {xml_file}")
        arithmos_teloneiou = ref_num_parts[2].strip()
        print(f"Extracted αριθμος τελωνειου: {arithmos_teloneiou}")

        # Determine if the file is "εισαγωγες" or "εξαγωγες"
        if root.find('.//NamEX17') is not None:
            type_file = "εξαγωγες"
        elif root.find('.//NamCE17') is not None:
            type_file = "εισαγωγες"
        else:
            raise ValueError(f"Could not determine the file type for XML file: {xml_file}")
        print(f"Determined file type: {type_file}")

        return teloneio, arithmos_teloneiou, type_file
    except ET.ParseError as e:
        print(f"Error parsing XML file {xml_file}: {e}")
        raise
    except Exception as e:
        print(f"An error occurred while extracting additional data from XML file {xml_file}: {e}")
        raise


def write_data_to_excel(excel_file, xml_filename, name, teloneio, arithmos_teloneiou, type_file):
    try:
        if not os.path.isfile(excel_file):
            raise FileNotFoundError(f"Excel file not found: {excel_file}")

        print(f"Opening Excel file: {excel_file}")
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        print(f"Writing data to Excel file: {excel_file}")
        # Find the next empty row
        next_row = sheet.max_row + 1

        # Remove the file extension from xml_filename
        xml_filename_no_ext = os.path.splitext(xml_filename)[0]

        # Write data to specific columns, column A is "column=1
        #                                 column B is "column=2
        #                                 etc
        # if columns are gonna change, please just modify only the "column" and nothing else
        #                      this "column"!
        #                          |
        #                          |
        #                          |
        #                          V
        sheet.cell(row=next_row, column=2, value=name)  # Column B (το ονομα του πελατη)
        sheet.cell(row=next_row, column=4, value=xml_filename_no_ext)  # Column D (το ονομα του αρχειου)
        sheet.cell(row=next_row, column=6, value=teloneio)  # Column F (τελωνειο)
        sheet.cell(row=next_row, column=3, value=arithmos_teloneiou)  # Column C (αριθμος υποθεσης)
        sheet.cell(row=next_row, column=11, value=type_file)  # Column K (εισαγωγη ή εξαγωγη)

        workbook.save(excel_file)
        print(f"Data successfully written to {excel_file}")
    except Exception as e:
        print(f"An error occurred while writing data to the Excel file: {e}")
        raise


def move_pdf_based_on_name(pdf_file, target_directory_base, name):
    try:
        print(f"Preparing to move PDF file: {pdf_file}")
        if not os.path.isfile(pdf_file):
            raise FileNotFoundError(f"PDF file not found: {pdf_file}")

        current_year = datetime.now().year
        target_path = os.path.join(target_directory_base, "ΠΕΛΑΤΕΣ", name, "ΔΙΑΣΑΦΗΣΕΙΣ", str(current_year))

        if not os.path.exists(target_path):
            print(f"Creating target directory: {target_path}")
            os.makedirs(target_path)

        print(f"Moving PDF to {target_path}")
        shutil.move(pdf_file, os.path.join(target_path, os.path.basename(pdf_file)))
        print(f"Successfully moved PDF to {target_path}")
    except Exception as e:
        print(f"An error occurred while moving the PDF file: {e}")
        raise


def process_files_in_directory(directory, target_directory_base, excel_file):
    if not os.path.exists(directory):
        print(f"Directory does not exist: {directory}")
        return

    files = os.listdir(directory)
    if not files:
        print(f"No files found in directory: {directory}")
        return

    print(f"Files found in directory: {files}")

    for filename in files:
        if filename.lower().endswith(".xml"):
            xml_file = os.path.join(directory, filename)
            pdf_file = os.path.join(directory, filename.rsplit('.', 1)[0] + ".pdf")

            print(f"Processing file: {filename}")

            try:
                name = get_name_from_xml(xml_file)
                teloneio, arithmos_teloneiou, type_file = extract_additional_data_from_xml(xml_file)
                write_data_to_excel(excel_file, filename, name, teloneio, arithmos_teloneiou, type_file)
                move_pdf_based_on_name(pdf_file, target_directory_base, name)
            except FileNotFoundError as e:
                print(f"PDF file not found for {filename}: {e}")
            except Exception as e:
                print(f"Failed to process {filename}: {e}")


def main():
    print("Opening file dialog...")
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    directory = filedialog.askdirectory(title="Select the directory containing XML and PDF files")

    if not directory:
        print("No directory selected. Exiting.")
        return

    print(f"Selected directory: {directory}")
    target_directory_base = "Z:/Public"
    excel_file = "D:/Downloads/arxeio.xlsx"  # Update with the path to your Excel file!!!

    process_files_in_directory(directory, target_directory_base, excel_file)


if __name__ == "__main__":
    main()
